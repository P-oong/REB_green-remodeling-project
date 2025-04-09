import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from copy import copy

def copy_sheet_with_styles(source_sheet, target_sheet, file_name):
    """시트의 모든 내용과 스타일, 병합 셀, 계산식 등을 복사"""
    
    # 열 너비 복사
    for i, column in enumerate(source_sheet.columns, 1):
        column_letter = get_column_letter(i)
        if source_sheet.column_dimensions[column_letter].width is not None:
            target_sheet.column_dimensions[column_letter].width = source_sheet.column_dimensions[column_letter].width
    
    # 행 높이 복사
    for i, row in enumerate(source_sheet.rows, 1):
        if source_sheet.row_dimensions[i].height is not None:
            target_sheet.row_dimensions[i].height = source_sheet.row_dimensions[i].height
    
    # 병합 셀 복사
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))
    
    # 셀 내용, 스타일, 계산식 복사
    for row in source_sheet.rows:
        for cell in row:
            target_cell = target_sheet.cell(row=cell.row, column=cell.column)
            
            # 값 복사 (계산식 포함)
            if cell.value is not None:
                target_cell.value = cell.value
            
            # 계산식이 있는 경우, 참조 시트 이름 변경 시도
            if cell.data_type == 'f':
                formula = cell.value
                
                # '원가계산서' -> '파일명(원가계산서)' 로 치환
                new_cost_sheet_name = f"{file_name}(원가계산서)"
                formula = formula.replace("'원가계산서'!", f"'{new_cost_sheet_name}'!")
                formula = formula.replace("원가계산서!", f"'{new_cost_sheet_name}'!") # 따옴표 없는 경우

                # '내역서' -> '파일명(내역서)' 로 치환
                new_statement_sheet_name = f"{file_name}(내역서)"
                formula = formula.replace("'내역서'!", f"'{new_statement_sheet_name}'!")
                formula = formula.replace("내역서!", f"'{new_statement_sheet_name}'!")

                target_cell.value = formula # 수정된 계산식 저장
                target_cell.data_type = 'f' # 타입 유지
            
            # 스타일 복사
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

def extract_number_prefix(filename):
    """파일 이름에서 숫자 접두사를 추출하여 정렬 키로 사용"""
    match = re.match(r'^(\d+)_', filename)
    if match:
        return int(match.group(1))
    return float('inf')  # 숫자 접두사가 없는 경우 가장 뒤로 정렬

def sort_sheets_in_excel(excel_file):
    """엑셀 파일 내의 시트를 01_부터 44_까지 정렬"""
    print(f"'{excel_file}' 파일의 시트를 정렬합니다...")
    
    # 엑셀 파일 로드
    wb = load_workbook(excel_file)
    
    # 시트 이름 목록 가져오기
    sheet_names = wb.sheetnames
    
    # 시트 이름에서 숫자 접두사 추출하여 정렬
    sorted_sheets = []
    for sheet_name in sheet_names:
        # 시트 이름에서 파일 번호 추출 (예: "01_파일명(원가계산서)" -> "01")
        match = re.match(r'^(\d+)_', sheet_name)
        if match:
            file_number = int(match.group(1))
            # 시트 유형 확인 (원가계산서는 1, 내역서는 2로 설정하여 원가계산서가 먼저 오도록)
            sheet_type = 1 if "(원가계산서)" in sheet_name else 2
            sorted_sheets.append((sheet_name, file_number, sheet_type))
        else:
            # 숫자 접두사가 없는 시트는 가장 뒤로
            sorted_sheets.append((sheet_name, float('inf'), 3))
    
    # 정렬: 1. 파일 번호순, 2. 원가계산서 먼저, 내역서 나중에
    sorted_sheets.sort(key=lambda x: (x[1], x[2]))
    
    # 정렬된 시트 이름 목록
    ordered_sheet_names = [sheet[0] for sheet in sorted_sheets]
    
    # 시트 순서 재배치
    for i, sheet_name in enumerate(ordered_sheet_names):
        # 시트 인덱스는 0부터 시작
        wb.move_sheet(sheet_name, offset=-len(wb.sheetnames) + i)
    
    # 결과 저장
    wb.save(excel_file)
    print(f"시트 정렬이 완료되었습니다. 총 {len(ordered_sheet_names)}개의 시트가 정렬되었습니다.")

def main():
    # 현재 디렉토리 경로
    current_dir = os.getcwd()
    
    # 결과를 저장할 엑셀 워크북 생성
    result_wb = Workbook()
    # 기본 시트 제거
    result_wb.remove(result_wb.active)
    
    # 디렉토리 내 모든 엑셀 파일 가져오기
    excel_files = [f for f in os.listdir(current_dir) if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$') and f != '● 내역서.xlsx']
    
    # 파일 이름의 숫자 접두사를 기준으로 정렬
    excel_files.sort(key=extract_number_prefix)
    
    print(f"총 {len(excel_files)}개의 엑셀 파일을 찾았습니다.")
    print(f"정렬된 순서로 처리합니다: {', '.join(excel_files[:3])}... 등")
    
    # 각 엑셀 파일 처리
    all_sheets = []  # 모든 시트 이름과 정렬 순서를 저장할 리스트
    
    for file in excel_files:
        file_path = os.path.join(current_dir, file)
        file_name = os.path.splitext(file)[0]  # 확장자 제외한 파일명
        
        try:
            print(f"처리 중: {file}")
            
            # 엑셀 파일 읽기 (openpyxl 사용)
            source_wb = load_workbook(file_path, data_only=False)  # data_only=False로 설정하여 계산식 유지
            
            # 원가계산서와 내역서 시트가 있는지 확인
            sheet_names = source_wb.sheetnames
            
            if "원가계산서" in sheet_names and "내역서" in sheet_names:
                # 새 시트 이름 생성
                cost_sheet_name = f"{file_name}(원가계산서)"
                statement_sheet_name = f"{file_name}(내역서)"
                
                # 결과 워크북에 시트 추가 - 원가계산서
                result_wb.create_sheet(cost_sheet_name)
                copy_sheet_with_styles(source_wb["원가계산서"], result_wb[cost_sheet_name], file_name)
                
                # 결과 워크북에 시트 추가 - 내역서
                result_wb.create_sheet(statement_sheet_name)
                copy_sheet_with_styles(source_wb["내역서"], result_wb[statement_sheet_name], file_name)
                
                # 시트 이름과 정렬 순서 저장
                sort_key = extract_number_prefix(file)
                all_sheets.append((cost_sheet_name, sort_key, 1))  # 원가계산서가 먼저 오도록 1 추가
                all_sheets.append((statement_sheet_name, sort_key, 2))  # 내역서는 나중에 오도록 2 추가
                
                print(f"  - '{cost_sheet_name}'와 '{statement_sheet_name}' 시트를 추가했습니다.")
            else:
                missing_sheets = []
                if "원가계산서" not in sheet_names:
                    missing_sheets.append("원가계산서")
                if "내역서" not in sheet_names:
                    missing_sheets.append("내역서")
                print(f"  - 경고: {file}에서 {', '.join(missing_sheets)} 시트를 찾을 수 없습니다.")
                
            source_wb.close()
                
        except Exception as e:
            print(f"  - 오류: {file} 처리 중 문제가 발생했습니다. 오류: {str(e)}")
    
    # 시트 순서 재정렬
    # 정렬 순서: 1. 파일 번호순, 2. 원가계산서 먼저, 내역서 나중에
    all_sheets.sort(key=lambda x: (x[1], x[2]))
    
    # 시트 순서 재배치
    ordered_sheet_names = [sheet[0] for sheet in all_sheets]
    
    # 시트를 정렬된 순서로 재배치 (openpyxl에서는 시트 위치 이동을 직접 지원하지 않아
    # 위치 매개변수를 사용하여 정렬된 순서대로 재배치)
    for i, sheet_name in enumerate(ordered_sheet_names):
        # 시트 인덱스는 0부터 시작
        result_wb.move_sheet(sheet_name, offset=-len(result_wb.sheetnames) + i)
    
    # 결과 저장
    result_wb.save('● 내역서.xlsx')
    print("\n통합이 완료되었습니다. '● 내역서.xlsx' 파일이 생성되었습니다.")
    print(f"총 {len(ordered_sheet_names)}개의 시트가 숫자순으로 정렬되어 저장되었습니다.")

if __name__ == "__main__":
    # 이미 생성된 엑셀 파일이 있는 경우 정렬만 수행
    if os.path.exists('● 내역서.xlsx'):
        sort_sheets_in_excel('● 내역서.xlsx')
    else:
        # 새로 통합 파일 생성
        main() 